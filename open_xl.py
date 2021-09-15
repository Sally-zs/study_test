from openpyxl import load_workbook

class TestCase:
    # 用于存放数据
    pass


class HandleExcel:

    def __init__(self, filename, sheetname=None):
        self.filename = filename  # 文件路径
        self.sheetname = sheetname  # 表单名

    def read_excel(self):
        '''
        读数据
        :return:
        '''
        wb = load_workbook(self.filename)  # 记载工作簿
        if self.sheetname is None:
            ws = wb.active  # 加载工作表
        else:
            ws = wb[self.sheetname]  # 加载指定工作表

        testcase_list = []  # 存放对象
        test_headers = []  # 存放表头信息
        for row in range(1, ws.max_row + 1):
            one_testcase = TestCase()  # 创建用例对象
            for column in range(1, ws.max_column + 1):
                one_cell_value = ws.cell(row, column).value
                if row == 1:
                    test_headers.append(str(one_cell_value))  # 获取表头字符串数据，方便后面调用
                else:
                    key = test_headers[column - 1]  # 获取表头字符串数据
                    if key == 'actual':
                        setattr(one_testcase, "actual_column", column)  # 动态创建列的值
                    elif key == 'result':
                        setattr(one_testcase, 'result_column', column)  # 动态创建列的值
                    setattr(one_testcase, key, one_cell_value)  # 动态创建key并赋值
            if row != 1:
                setattr(one_testcase, 'row', row)  # 动态创建每一行
                testcase_list.append(one_testcase)  # 将每个对象添加到列表

        return testcase_list

    def write_excel(self, one_testcase, actual_value, result):
        '''
        写操作
        :param one_testcase:
        :param actual_value: 写入值
        :param result:写入结果
        :return:
        '''
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        ws.cell(one_testcase.row, one_testcase.actual_column, value=actual_value)  # 写入操作
        ws.cell(one_testcase.row, one_testcase.result_column, value=result)  # 写入操作
        wb.save(self.filename)  # 保存文件


if __name__ == '__main__':
    excel_filename = r"D:\ivy_person\study_test\testcase_data.xlsx"  # 文件路径，xlsx文件
    sheet_name = "user"
    do_excel = HandleExcel(excel_filename, sheet_name)
    testcases_data = do_excel.read_excel()
    do_excel.write_excel(testcases_data[0], '10', '20')
    pass
print("11111")